import os
import smtplib
import random
import string
import configparser
import tempfile
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import win32com.client

# --- Configuration ---
CONFIG_FILE = 'config.ini'
MOBILE_TEMPLATE = 'Mobile_Bill_Template.xlsx'
LANDLINE_TEMPLATE = 'Landline_Bill_Template.xlsx'

def generate_random_bill_no():
    letters = ''.join(random.choices(string.ascii_uppercase, k=2))
    digits = ''.join(random.choices(string.digits, k=14))
    return f"{letters}{digits}"

def update_excel_file(template_path, temp_dir, is_mobile_bill):
    print(f"  Loading template: {os.path.basename(template_path)}")
    try:
        wb = load_workbook(filename=template_path)
        ws = wb.active
    except FileNotFoundError:
        return None, f"Template file not found: {os.path.basename(template_path)}"

    today = datetime.now()
    statement_date_dt = today.replace(day=23) - relativedelta(months=1)
    period_start_dt = today.replace(day=23) - relativedelta(months=2)
    period_end_dt = today.replace(day=22) - relativedelta(months=1)
    due_date_dt = today.replace(day=12)

    statement_date_str = f"Statement Date:{statement_date_dt.strftime('%d %b %Y')}"
    period_start_str = period_start_dt.strftime('%d %b %Y')
    period_end_str = period_end_dt.strftime('%d %b %Y')
    statement_period_str = f"Statement Period:{period_start_str}-{period_end_str}"
    due_date_q7_str = due_date_dt.strftime('%d-%b-%Y')
    due_date_s12_str = f"Amount after due date ({due_date_dt.strftime('%d %B')})"
    new_bill_no = f"Bill No. {generate_random_bill_no()}"

    if is_mobile_bill:
        ws['J5'] = statement_date_str
        ws['J6'] = statement_period_str
        ws['Q7'] = due_date_q7_str
        ws['S12'] = due_date_s12_str
        ws['H82'] = new_bill_no
        temp_filename = "temp_mobile_bill.xlsx"
    else:
        ws['J7'] = statement_date_str
        ws['J8'] = statement_period_str
        ws['Q7'] = due_date_q7_str
        ws['S12'] = due_date_s12_str
        ws['H82'] = new_bill_no
        temp_filename = "temp_landline_bill.xlsx"

    temp_excel_path = os.path.join(temp_dir, temp_filename)
    wb.save(temp_excel_path)
    print(f"  Generated: {temp_filename}")
    return temp_excel_path, None

def convert_excel_to_pdf(excel_path, pdf_path):
    print(f"  Converting to PDF...")
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
        workbook.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        workbook.Close(False)
        print(f"  PDF created: {os.path.basename(pdf_path)}")
        return True, None
    except Exception as e:
        return False, f"Could not convert to PDF: {e}"
    finally:
        if excel:
            excel.Quit()

def send_email(config, attachments):
    print("\n3. Sending email...")
    try:
        sender_email = config.get('Email', 'SENDER_EMAIL')
        sender_password = config.get('Email', 'SENDER_PASSWORD')
        recipient_email = config.get('Email', 'RECIPIENT_EMAIL')
        smtp_server = config.get('Email', 'SMTP_SERVER')
        smtp_port = int(config.get('Email', 'SMTP_PORT'))
    except (configparser.NoSectionError, configparser.NoOptionError) as e:
        return False, f"Could not read setting '{e.option}' from config.ini.", None

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = f"Your Bills for {datetime.now().strftime('%B %Y')}"
    msg.attach(MIMEText("Please find your monthly bills attached.\n\nThank you.", 'plain'))

    for file_path in attachments:
        with open(file_path, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=os.path.basename(file_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            msg.attach(part)
        print(f"  Attached: {os.path.basename(file_path)}")

    try:
        print(f"  Connecting to {smtp_server}:{smtp_port}...")
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        print(f"  Email sent successfully to {recipient_email}")
        return True, None, recipient_email
    except Exception as e:
        return False, f"Failed to send email: {e}", recipient_email

def main():
    print("="*60)
    print("BILL GENERATION SCRIPT")
    print("="*60)

    temp_dir = tempfile.gettempdir()
    files_to_cleanup = []

    try:
        # Load config
        print("\nLoading configuration...")
        config = configparser.ConfigParser()
        config.read(CONFIG_FILE)
        print(f"  Config loaded from {CONFIG_FILE}")

        # --- Process Mobile Bill ---
        print("\n1. Generating Mobile Bill...")
        mobile_temp_xlsx, error = update_excel_file(MOBILE_TEMPLATE, temp_dir, is_mobile_bill=True)
        if error:
            raise Exception(error)
        files_to_cleanup.append(mobile_temp_xlsx)

        mobile_pdf_name = f"Mobile Bill {datetime.now().strftime('%B-%y')}.pdf"
        mobile_pdf_path = os.path.join(temp_dir, mobile_pdf_name)
        files_to_cleanup.append(mobile_pdf_path)
        success, error = convert_excel_to_pdf(mobile_temp_xlsx, mobile_pdf_path)
        if not success:
            raise Exception(error)

        # --- Process Landline Bill ---
        print("\n2. Generating Landline Bill...")
        landline_temp_xlsx, error = update_excel_file(LANDLINE_TEMPLATE, temp_dir, is_mobile_bill=False)
        if error:
            raise Exception(error)
        files_to_cleanup.append(landline_temp_xlsx)

        landline_pdf_name = f"Landline Bill {datetime.now().strftime('%B-%y')}.pdf"
        landline_pdf_path = os.path.join(temp_dir, landline_pdf_name)
        files_to_cleanup.append(landline_pdf_path)
        success, error = convert_excel_to_pdf(landline_temp_xlsx, landline_pdf_path)
        if not success:
            raise Exception(error)

        # --- Send Email ---
        pdfs_to_send = [mobile_pdf_path, landline_pdf_path]
        success, error, recipient = send_email(config, pdfs_to_send)
        if not success:
            raise Exception(error)

        print("\n" + "="*60)
        print("SUCCESS!")
        print(f"Bills sent to: {recipient}")
        print("="*60)

    except Exception as e:
        print("\n" + "="*60)
        print("ERROR OCCURRED")
        print(f"{e}")
        print("="*60)

    finally:
        # Cleanup
        print("\nCleaning up temporary files...")
        for f in files_to_cleanup:
            if os.path.exists(f):
                try:
                    os.remove(f)
                    print(f"  Deleted: {os.path.basename(f)}")
                except Exception as e:
                    print(f"  Warning: Could not delete {os.path.basename(f)}")

if __name__ == "__main__":
    main()
