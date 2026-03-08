#!/usr/bin/env python3
"""
Monthly Bill Generator (GitHub Actions)

Uses Excel templates + LibreOffice (Linux) or Excel COM (Windows).
"""

import os
import sys
import subprocess
import tempfile
import shutil
from datetime import datetime

from bill_utils import compute_billing_dates, send_email_smtp, pdf_filename

# Templates
MOBILE_TEMPLATE = 'Mobile_Bill_Template.xlsx'
LANDLINE_TEMPLATE = 'Landline_Bill_Template.xlsx'


def get_env(name):
    """Get required environment variable."""
    value = os.environ.get(name)
    if not value:
        print(f"ERROR: Missing environment variable: {name}")
        sys.exit(1)
    return value


def update_excel(template_path, output_xlsx, is_mobile, dates):
    """Update Excel file with new dates using openpyxl."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(template_path)
        ws = wb.active
        
        if is_mobile:
            ws['J5'] = dates['statement_date_str']
            ws['J6'] = dates['statement_period_str']
        else:
            ws['J7'] = dates['statement_date_str']
            ws['J8'] = dates['statement_period_str']
        
        ws['Q7'] = dates['due_date_q7_str']
        ws['S12'] = dates['due_date_s12_str']
        ws['H82'] = dates['bill_no_str']
        
        wb.save(output_xlsx)
        return True
    except Exception as e:
        print(f"Error updating Excel: {e}")
        return False


def convert_to_pdf(excel_path, pdf_path):
    """Convert Excel to PDF using LibreOffice (Linux)."""
    result = subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'pdf',
         '--outdir', os.path.dirname(excel_path), excel_path],
        capture_output=True, text=True, timeout=120
    )
    
    if result.returncode != 0:
        raise RuntimeError(f"PDF conversion failed: {result.stderr}")
    
    # Rename output
    base = os.path.splitext(os.path.basename(excel_path))[0]
    generated = os.path.join(os.path.dirname(excel_path), base + '.pdf')
    
    if os.path.exists(generated):
        shutil.move(generated, pdf_path)
    else:
        raise RuntimeError(f"PDF not created")
    
    return pdf_path


def main():
    print("=" * 60)
    print("MONTHLY BILL GENERATOR")
    print(f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Get credentials
    sender_email = get_env('SENDER_EMAIL')
    sender_password = get_env('SENDER_PASSWORD')
    recipient_email = get_env('RECIPIENT_EMAIL')
    smtp_server = get_env('SMTP_SERVER')
    smtp_port = int(get_env('SMTP_PORT'))

    dates = compute_billing_dates()
    print(f"\nBilling dates: {dates['statement_date_str']}")

    temp_dir = tempfile.mkdtemp()
    files_to_cleanup = []

    try:
        # Mobile Bill
        print("\n1. Generating Mobile Bill...")
        mobile_xlsx = os.path.join(temp_dir, 'mobile.xlsx')
        mobile_pdf = os.path.join(temp_dir, pdf_filename("Mobile"))
        
        update_excel(MOBILE_TEMPLATE, mobile_xlsx, True, dates)
        convert_to_pdf(mobile_xlsx, mobile_pdf)
        files_to_cleanup.extend([mobile_xlsx, mobile_pdf])
        print(f"   Generated: {os.path.basename(mobile_pdf)}")

        # Landline Bill
        print("\n2. Generating Landline Bill...")
        landline_xlsx = os.path.join(temp_dir, 'landline.xlsx')
        landline_pdf = os.path.join(temp_dir, pdf_filename("Landline"))
        
        update_excel(LANDLINE_TEMPLATE, landline_xlsx, False, dates)
        convert_to_pdf(landline_xlsx, landline_pdf)
        files_to_cleanup.extend([landline_xlsx, landline_pdf])
        print(f"   Generated: {os.path.basename(landline_pdf)}")

        # Send email
        print("\n3. Sending email...")
        attachments = [mobile_pdf, landline_pdf]
        for a in attachments:
            print(f"   Attached: {os.path.basename(a)}")

        success, error, _ = send_email_smtp(
            sender_email, sender_password, recipient_email,
            smtp_server, smtp_port, attachments
        )
        if not success:
            raise RuntimeError(error)
        print(f"   Email sent successfully to {recipient_email}")

        print("\n" + "=" * 60)
        print("SUCCESS!")
        print("=" * 60)

    except Exception as e:
        print("\n" + "=" * 60)
        print(f"ERROR: {e}")
        print("=" * 60)
        sys.exit(1)

    finally:
        print("\nCleaning up...")
        for f in files_to_cleanup:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except:
                    pass


if __name__ == "__main__":
    main()
