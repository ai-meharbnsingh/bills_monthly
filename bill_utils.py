"""
Shared utilities for bill generation.

Contains common functions used by main.py (Linux/Docker), generate_bills.py (Windows GUI),
and run_bills_console.py (Windows console).
"""

import os
import random
import string
import smtplib
import tempfile
import shutil
from datetime import datetime
from dateutil.relativedelta import relativedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Optional: xlwings for Excel editing with Microsoft Excel (Windows/Mac)
USE_XLWINGS = False
try:
    import xlwings as xw
    try:
        _app = xw.App(visible=False)
        _app.quit()
        USE_XLWINGS = True
    except:
        pass
except ImportError:
    pass

# Optional: openpyxl for Excel editing without Excel (Linux)
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def generate_random_bill_no():
    """Generate a random bill number: 2 uppercase letters + 14 digits."""
    letters = ''.join(random.choices(string.ascii_uppercase, k=2))
    digits = ''.join(random.choices(string.digits, k=14))
    return f"{letters}{digits}"


def compute_billing_dates():
    """Compute all billing dates relative to today.

    Returns a dict with keys:
        statement_date_str, statement_period_str,
        due_date_q7_str, due_date_s12_str, bill_no_str
    """
    today = datetime.now()
    statement_date_dt = today.replace(day=23) - relativedelta(months=1)
    period_start_dt = today.replace(day=23) - relativedelta(months=2)
    period_end_dt = today.replace(day=22) - relativedelta(months=1)
    due_date_dt = today.replace(day=12)

    return {
        'statement_date_str': f"Statement Date:{statement_date_dt.strftime('%d %b %Y')}",
        'statement_period_str': (
            f"Statement Period:{period_start_dt.strftime('%d %b %Y')}"
            f"-{period_end_dt.strftime('%d %b %Y')}"
        ),
        'due_date_q7_str': due_date_dt.strftime('%d-%b-%Y'),
        'due_date_s12_str': f"Amount after due date ({due_date_dt.strftime('%d %B')})",
        'bill_no_str': f"Bill No. {generate_random_bill_no()}",
    }


def update_excel_file(template_path, temp_dir, is_mobile_bill):
    """Update an Excel template with current billing dates and a random bill number.

    Args:
        template_path: Path to the .xlsx template.
        temp_dir: Directory to write the temporary file into.
        is_mobile_bill: True for mobile, False for landline.

    Returns:
        (temp_excel_path, error_message) — error_message is None on success.
    """
    if not os.path.exists(template_path):
        return None, f"Template file not found: {os.path.basename(template_path)}"

    dates = compute_billing_dates()

    if USE_XLWINGS:
        # Use xlwings (preserves images and formatting via Excel app)
        try:
            app = xw.App(visible=False)
            wb = app.books.open(template_path)
            ws = wb.sheets[0]

            if is_mobile_bill:
                ws.range('J5').value = dates['statement_date_str']
                ws.range('J6').value = dates['statement_period_str']
                temp_filename = "temp_mobile_bill.xlsx"
            else:
                ws.range('J7').value = dates['statement_date_str']
                ws.range('J8').value = dates['statement_period_str']
                temp_filename = "temp_landline_bill.xlsx"

            # Common cells for both bill types
            ws.range('Q7').value = dates['due_date_q7_str']
            ws.range('S12').value = dates['due_date_s12_str']
            ws.range('H82').value = dates['bill_no_str']

            temp_excel_path = os.path.join(temp_dir, temp_filename)
            wb.save(temp_excel_path)
            wb.close()
            app.quit()
            return temp_excel_path, None
        except Exception as e:
            try:
                app.quit()
            except:
                pass
            return None, f"xlwings error: {e}"
    elif load_workbook is not None:
        # Fallback to openpyxl (images will be lost - for GitHub Actions only)
        try:
            wb = load_workbook(filename=template_path)
            ws = wb.active
        except FileNotFoundError:
            return None, f"Template file not found: {os.path.basename(template_path)}"

        if is_mobile_bill:
            ws['J5'] = dates['statement_date_str']
            ws['J6'] = dates['statement_period_str']
            temp_filename = "temp_mobile_bill.xlsx"
        else:
            ws['J7'] = dates['statement_date_str']
            ws['J8'] = dates['statement_period_str']
            temp_filename = "temp_landline_bill.xlsx"

        # Common cells for both bill types
        ws['Q7'] = dates['due_date_q7_str']
        ws['S12'] = dates['due_date_s12_str']
        ws['H82'] = dates['bill_no_str']

        temp_excel_path = os.path.join(temp_dir, temp_filename)
        wb.save(temp_excel_path)
        return temp_excel_path, None


def convert_excel_to_pdf(excel_path, pdf_path):
    """Convert Excel to PDF using xlwings (preserves all formatting)."""
    if not USE_XLWINGS:
        raise RuntimeError("PDF conversion requires xlwings/Excel app")
    
    app = xw.App(visible=False)
    try:
        wb = app.books.open(excel_path)
        wb.to_pdf(pdf_path)
        wb.close()
        app.quit()
    except Exception as e:
        try:
            app.quit()
        except:
            pass
        raise RuntimeError(f"PDF conversion failed: {e}")


def send_email_smtp(sender_email, sender_password, recipient_email,
                    smtp_server, smtp_port, attachments):
    """Send an email with PDF bill attachments.

    Args:
        sender_email: Sender's email address.
        sender_password: Sender's email password / app password.
        recipient_email: Recipient's email address.
        smtp_server: SMTP server hostname.
        smtp_port: SMTP port (465 for SSL, others use STARTTLS).
        attachments: List of file paths to attach.

    Returns:
        (success, error_message, recipient_email)
    """
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = f"Your Bills for {datetime.now().strftime('%B %Y')}"
    msg.attach(MIMEText(
        "Please find your monthly bills attached.\n\nThank you.", 'plain'
    ))

    for file_path in attachments:
        with open(file_path, 'rb') as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(file_path))
            part['Content-Disposition'] = (
                f'attachment; filename="{os.path.basename(file_path)}"'
            )
            msg.attach(part)

    try:
        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(sender_email, sender_password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.send_message(msg)
        return True, None, recipient_email
    except Exception as e:
        return False, f"Failed to send email: {e}", recipient_email


def pdf_filename(bill_type):
    """Generate a PDF filename like 'Mobile Bill February-26.pdf'."""
    return f"{bill_type} Bill {datetime.now().strftime('%B-%y')}.pdf"
