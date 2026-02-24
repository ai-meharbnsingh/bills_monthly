import os
import smtplib
import random
import string
import configparser
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import sys
import subprocess # For calling external programs
import tempfile # For creating files in a temporary directory
import threading # To keep the status window responsive
import time # For delays

# --- Global Configuration Constants ---
CONFIG_FILE = 'config.ini'
MOBILE_TEMPLATE = 'Mobile_Bill_Template.xlsx'
LANDLINE_TEMPLATE = 'Landline_Bill_Template.xlsx'

# --- Status Window Class ---
class StatusWindow:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Processing...")
        
        parent.update_idletasks()
        width = 350
        height = 100
        x = (parent.winfo_screenwidth() // 2) - (width // 2)
        y = (parent.winfo_screenheight() // 2) - (height // 2)
        self.top.geometry(f'{width}x{height}+{x}+{y}')
        
        self.label = tk.Label(self.top, text="Initializing...", padx=20, pady=20, font=("Helvetica", 10))
        self.label.pack()
        self.top.lift()

    def update_status(self, text):
        self.label.config(text=text)
        self.top.update()

    def close(self):
        self.top.destroy()

# --- Main Functions ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def generate_random_bill_no():
    letters = ''.join(random.choices(string.ascii_uppercase, k=2))
    digits = ''.join(random.choices(string.digits, k=14))
    return f"{letters}{digits}"

def update_excel_file(template_path, temp_dir, is_mobile_bill):
    try:
        wb = load_workbook(filename=template_path)
        ws = wb.active
    except FileNotFoundError:
        return None, f"Template file not found:\n{os.path.basename(template_path)}"

    today = datetime.now()
    statement_date_dt = today.replace(day=23) - relativedelta(months=1)
    period_start_dt = today.replace(day=23) - relativedelta(months=2)
    period_end_dt = today.replace(day=22) - relativedelta(months=1)
    due_date_dt = today.replace(day=12)

    statement_date_str = f"Statement Date:{statement_date_dt.strftime('%d %b %Y')}"
    period_start_str = period_start_dt.strftime('%d %b %Y')
    period_end_str = period_end_dt.strftime('%d %b %Y')
    statement_period_str = f"Statement Period:{period_start_str}-{period_end_str}"
    due_date_q7_str = due_date_dt.strftime('%d-%b-%Y')
    due_date_s12_str = f"Amount after due date ({due_date_dt.strftime('%d %B')})"
    new_bill_no = f"Bill No. {generate_random_bill_no()}"

    if is_mobile_bill:
        ws['J5'] = statement_date_str; ws['J6'] = statement_period_str
        ws['Q7'] = due_date_q7_str; ws['S12'] = due_date_s12_str; ws['H82'] = new_bill_no
        temp_filename = "temp_mobile_bill.xlsx"
    else:
        ws['J7'] = statement_date_str; ws['J8'] = statement_period_str
        ws['Q7'] = due_date_q7_str; ws['S12'] = due_date_s12_str; ws['H82'] = new_bill_no
        temp_filename = "temp_landline_bill.xlsx"
    
    temp_excel_path = os.path.join(temp_dir, temp_filename)
    wb.save(temp_excel_path)
    return temp_excel_path, None

def convert_excel_to_pdf(excel_path, pdf_path):
    """
    Convert Excel to PDF using VBScript - works without COM issues
    """
    try:
        # Convert to absolute paths
        abs_excel_path = os.path.abspath(excel_path)
        abs_pdf_path = os.path.abspath(pdf_path)

        # Create VBScript to convert Excel to PDF
        vbs_script = f'''
Set objExcel = CreateObject("Excel.Application")
objExcel.Visible = False
objExcel.DisplayAlerts = False

Set objWorkbook = objExcel.Workbooks.Open("{abs_excel_path}")
objWorkbook.ExportAsFixedFormat 0, "{abs_pdf_path}"
objWorkbook.Close False

objExcel.Quit
Set objWorkbook = Nothing
Set objExcel = Nothing
'''

        # Save VBScript to temp file
        vbs_path = os.path.join(tempfile.gettempdir(), "excel_to_pdf.vbs")
        with open(vbs_path, 'w') as f:
            f.write(vbs_script)

        # Run VBScript
        result = subprocess.run(['cscript', '//nologo', vbs_path],
                              capture_output=True,
                              text=True,
                              timeout=30)

        # Clean up VBScript
        try:
            os.remove(vbs_path)
        except:
            pass

        # Check if PDF was created
        if os.path.exists(abs_pdf_path):
            return True, None
        else:
            error_msg = result.stderr if result.stderr else "Unknown error"
            return False, f"Could not convert to PDF.\nPlease ensure MS Excel is installed.\n\nError: {error_msg}"

    except subprocess.TimeoutExpired:
        return False, "Excel conversion timed out. Please close any Excel windows and try again."
    except Exception as e:
        return False, f"Could not convert to PDF.\nPlease ensure MS Excel is installed.\n\nError: {e}"

def send_email(config, attachments):
    try:
        sender_email = config.get('Email', 'SENDER_EMAIL')
        sender_password = config.get('Email', 'SENDER_PASSWORD')
        recipient_email = config.get('Email', 'RECIPIENT_EMAIL')
        smtp_server = config.get('Email', 'SMTP_SERVER')
        smtp_port = int(config.get('Email', 'SMTP_PORT'))
    except (configparser.NoSectionError, configparser.NoOptionError) as e:
        return False, f"Could not read setting '{e.option}' from config.ini.", None

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = f"Your Bills for {datetime.now().strftime('%B %Y')}"
    msg.attach(MIMEText("Please find your monthly bills attached.\n\nThank you.", 'plain'))

    for file_path in attachments:
        with open(file_path, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=os.path.basename(file_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            msg.attach(part)
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        return True, None, recipient_email
    except Exception as e:
        return False, f"Failed to send email. Check config and internet.\n\nError: {e}", recipient_email

# We now pass the main tkinter 'root' window as an argument
def main_process(root, status_window, config_path, mobile_template_path, landline_template_path):
    temp_dir = tempfile.gettempdir()
    files_to_cleanup = []
    
    try:
        status_window.update_status("Loading configuration...")
        config = configparser.ConfigParser()
        config.read(config_path)

        # --- Process Mobile Bill ---
        status_window.update_status("Generating Mobile Bill...")
        mobile_temp_xlsx, error = update_excel_file(mobile_template_path, temp_dir, is_mobile_bill=True)
        if error: raise Exception(error)
        files_to_cleanup.append(mobile_temp_xlsx)

        status_window.update_status("Converting Mobile Bill to PDF...")
        mobile_pdf_name = f"Mobile Bill {datetime.now().strftime('%B-%y')}.pdf"
        mobile_pdf_path = os.path.join(temp_dir, mobile_pdf_name)
        files_to_cleanup.append(mobile_pdf_path)
        success, error = convert_excel_to_pdf(mobile_temp_xlsx, mobile_pdf_path)
        if not success: raise Exception(error)

        # --- Process Landline Bill ---
        status_window.update_status("Generating Landline Bill...")
        landline_temp_xlsx, error = update_excel_file(landline_template_path, temp_dir, is_mobile_bill=False)
        if error: raise Exception(error)
        files_to_cleanup.append(landline_temp_xlsx)
        
        status_window.update_status("Converting Landline Bill to PDF...")
        landline_pdf_name = f"Landline Bill {datetime.now().strftime('%B-%y')}.pdf"
        landline_pdf_path = os.path.join(temp_dir, landline_pdf_name)
        files_to_cleanup.append(landline_pdf_path)
        success, error = convert_excel_to_pdf(landline_temp_xlsx, landline_pdf_path)
        if not success: raise Exception(error)
        
        # --- Send Email ---
        status_window.update_status("Connecting to email server...")
        pdfs_to_send = [mobile_pdf_path, landline_pdf_path]
        success, error, recipient = send_email(config, pdfs_to_send)
        if not success: raise Exception(error)
        
        status_window.close()
        messagebox.showinfo("Success!", f"Bills have been successfully generated and sent to {recipient}")

    except Exception as e:
        status_window.close()
        messagebox.showerror("An Error Occurred", str(e))
    
    finally:
        # --- Cleanup ---
        for f in files_to_cleanup:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except Exception as e:
                    print(f"Warning: Could not delete temporary file {f}. Error: {e}")
        
        # **** THE CRITICAL FIX IS HERE ****
        # This tells the main GUI loop to exit cleanly.
        root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    status_win = StatusWindow(root)

    config_file_path = resource_path(CONFIG_FILE)
    mobile_template_file_path = resource_path(MOBILE_TEMPLATE)
    landline_template_file_path = resource_path(LANDLINE_TEMPLATE)

    # We now pass 'root' as the first argument
    main_thread = threading.Thread(
        target=main_process, 
        args=(root, status_win, config_file_path, mobile_template_file_path, landline_template_file_path)
    )
    main_thread.start()
    
    root.mainloop()